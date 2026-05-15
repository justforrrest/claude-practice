import openpyxl, json, sys

path = r'C:\Users\feelw\OneDrive\문서\★HOME PC★\claude-practice\약재재고\약재관리.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print("시트 목록:", wb.sheetnames)

# 약재관리 탭 찾기
sheet_name = None
for s in wb.sheetnames:
    if '약재관리' in s:
        sheet_name = s
        break
if not sheet_name:
    sheet_name = wb.sheetnames[0]

ws = wb[sheet_name]
print(f"\n사용 시트: {sheet_name}")
print(f"최대 행: {ws.max_row}, 최대 열: {ws.max_column}")

# F~I열 (6~9열) 헤더 + 데이터 확인
print("\n--- F~I 열 데이터 (첫 30행) ---")
for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=12, values_only=True):
    if any(v is not None for v in row):
        print(row)
